from operator import truediv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

class Excel:
    def __init__(self):
        self.path = "./orders.xlsx"
        self.wb = load_workbook(self.path)
        self.sheet = self.wb.get_sheet_by_name('orders')
        self.max_row = self.sheet.max_row
        self.max_column = 0

    def check_spam(self, id):
        for i in range(1, self.max_row + 1):
            if self.sheet["A" + str(i)].value == id:
                return True
        return False

    def do_excel(self, info):
        self.max_row += 1
        for i in info:
            self.max_column += 1
            cell = self.sheet.cell(row = self.max_row, column = self.max_column)
            cell.value = i
        self.wb.save(self.path)
    
    def edit(self, id, column, value):
        for i in range(1, self.max_row + 1):
            if self.sheet["A" + str(i)].value == id:
                self.sheet[column + str(i)].value = value
                self.wb.save(self.path)
                break
    
    def delete(self, id):
        for i in range(1, self.max_row + 1):
            if str(self.sheet["A" + str(i)].value) == id:
                self.sheet.delete_rows(i, 1)
                self.wb.save(self.path)
                break

    def show_order(self, id):
        for i in range(1, self.max_row + 1):
            if self.sheet["A" + str(i)].value == id:
                return f"Почта: {self.sheet['B' + str(i)].value} \nФИО: {self.sheet['C' + str(i)].value} \nТелефон: {self.sheet['D' + str(i)].value} \nКол-во: {self.sheet['E' + str(i)].value} \nМесто: {self.sheet['F' + str(i)].value} \nДата: {self.sheet['G' + str(i)].value} \nКомментарий: {self.sheet['H' + str(i)].value}"